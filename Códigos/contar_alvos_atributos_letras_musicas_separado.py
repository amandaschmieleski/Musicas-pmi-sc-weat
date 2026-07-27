from __future__ import annotations

import csv
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


# ============================================================
# CAMINHOS
# ============================================================

PASTA = Path(r"D:\Downloads")

ARQUIVO_CSV = PASTA / "25_julho_merged_df_minusculo_sem_duplicadas.csv"
ARQUIVO_LISTAS = PASTA / "26_julho_Atributos_e_alvos_listas_Chen_Amanda_final.xlsx"

ARQUIVO_SAIDA = PASTA / "frequencia_alvos_atributos_letras_das_musicas.xlsx"

# A contagem será feita SOMENTE nesta coluna.
COLUNA_LETRA_ESPERADA = "Letra da Música"


# ============================================================
# FUNÇÕES DE NORMALIZAÇÃO E LEITURA
# ============================================================

def normalizar_cabecalho(valor: object) -> str:
    texto = "" if valor is None else str(valor)
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = texto.casefold().strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def normalizar_texto_busca(valor: object) -> str:
    texto = "" if valor is None else str(valor)
    texto = unicodedata.normalize("NFC", texto).casefold()
    texto = (
        texto.replace("’", "'")
        .replace("‘", "'")
        .replace("–", "-")
        .replace("—", "-")
        .replace("−", "-")
    )
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def detectar_csv(caminho: Path) -> tuple[str, str, list[str]]:
    encodings = ("utf-8-sig", "utf-8", "cp1252", "latin1")
    delimitadores = ",;\t|"

    ultimo_erro: Exception | None = None

    for encoding in encodings:
        try:
            with caminho.open("r", encoding=encoding, newline="") as arquivo:
                amostra = arquivo.read(200_000)

            if not amostra:
                raise ValueError("O CSV está vazio.")

            try:
                separador = csv.Sniffer().sniff(
                    amostra,
                    delimiters=delimitadores,
                ).delimiter
            except csv.Error:
                primeira_linha = amostra.splitlines()[0]
                separador = max(
                    delimitadores,
                    key=lambda delimitador: primeira_linha.count(delimitador),
                )

            cabecalho = pd.read_csv(
                caminho,
                sep=separador,
                encoding=encoding,
                nrows=0,
            )

            return encoding, separador, list(cabecalho.columns)

        except (UnicodeDecodeError, pd.errors.ParserError, ValueError) as erro:
            ultimo_erro = erro

    raise RuntimeError(
        f"Não foi possível identificar o formato do CSV. Último erro: {ultimo_erro}"
    )


def localizar_coluna_letra_musica(colunas: list[str]) -> str:
    procurada = normalizar_cabecalho(COLUNA_LETRA_ESPERADA)

    equivalencias = {
        normalizar_cabecalho(coluna): coluna
        for coluna in colunas
    }

    if procurada not in equivalencias:
        raise KeyError(
            "A coluna 'Letra da Música' não foi encontrada.\n"
            f"Colunas disponíveis: {colunas}"
        )

    return equivalencias[procurada]


def validar_colunas(
    dataframe: pd.DataFrame,
    colunas_obrigatorias: list[str],
    nome_aba: str,
) -> None:
    faltantes = [
        coluna
        for coluna in colunas_obrigatorias
        if coluna not in dataframe.columns
    ]

    if faltantes:
        raise KeyError(
            f"A aba '{nome_aba}' não possui as colunas: {faltantes}.\n"
            f"Colunas disponíveis: {list(dataframe.columns)}"
        )


# ============================================================
# LEITURA DAS LISTAS DE ALVOS E ATRIBUTOS
# ============================================================

def carregar_listas(caminho_xlsx: Path) -> pd.DataFrame:
    colunas = ["Tipo", "Categoria", "Termo no código", "Origem"]

    alvos = pd.read_excel(
        caminho_xlsx,
        sheet_name="Detalhe_alvos",
        dtype=str,
    )
    atributos = pd.read_excel(
        caminho_xlsx,
        sheet_name="Detalhe_atributos",
        dtype=str,
    )

    validar_colunas(alvos, colunas, "Detalhe_alvos")
    validar_colunas(atributos, colunas, "Detalhe_atributos")

    listas = pd.concat(
        [
            alvos[colunas],
            atributos[colunas],
        ],
        ignore_index=True,
    )

    listas = listas.rename(columns={"Termo no código": "Termo"})
    listas["Termo"] = listas["Termo"].fillna("").astype(str).str.strip()
    listas["Origem"] = listas["Origem"].fillna("").astype(str).str.strip()
    listas["Tipo"] = listas["Tipo"].fillna("").astype(str).str.strip()
    listas["Categoria"] = listas["Categoria"].fillna("").astype(str).str.strip()

    listas = listas[listas["Termo"] != ""].copy()
    listas["Termo_normalizado"] = listas["Termo"].map(normalizar_texto_busca)
    listas["Origem_normalizada"] = listas["Origem"].map(normalizar_texto_busca)

    listas = listas.drop_duplicates(
        subset=[
            "Tipo",
            "Categoria",
            "Origem",
            "Termo_normalizado",
        ]
    ).reset_index(drop=True)

    return listas


# ============================================================
# CONTAGEM SOMENTE EM "LETRA DA MÚSICA"
# ============================================================

def construir_regex(termos: list[str]) -> re.Pattern[str]:
    termos_unicos = sorted(
        {termo for termo in termos if termo},
        key=lambda termo: (-len(termo), termo),
    )

    if not termos_unicos:
        raise ValueError("A lista de termos está vazia.")

    alternativas = "|".join(re.escape(termo) for termo in termos_unicos)

    return re.compile(
        rf"(?<!\w)(?:{alternativas})(?!\w)"
    )


def contar_nas_letras(
    letras: pd.Series,
    termos: list[str],
) -> tuple[Counter[str], dict[str, set[int]]]:
    padrao = construir_regex(termos)

    frequencias: Counter[str] = Counter()
    musicas_por_termo: dict[str, set[int]] = defaultdict(set)

    total = len(letras)

    for indice, letra in enumerate(letras, start=1):
        encontrados = [
            correspondencia.group(0)
            for correspondencia in padrao.finditer(letra)
        ]

        if encontrados:
            frequencias.update(encontrados)

            for termo in set(encontrados):
                musicas_por_termo[termo].add(indice)

        if indice % 25_000 == 0 or indice == total:
            print(f"Processadas {indice:,} de {total:,} músicas...")

    return frequencias, musicas_por_termo


# ============================================================
# RESULTADOS
# ============================================================

def montar_frequencia_por_termo(
    listas: pd.DataFrame,
    frequencias: Counter[str],
    musicas_por_termo: dict[str, set[int]],
) -> pd.DataFrame:
    resultado = listas[
        ["Termo", "Tipo", "Categoria", "Origem", "Termo_normalizado"]
    ].copy()

    resultado["Frequência"] = resultado["Termo_normalizado"].map(
        lambda termo: int(frequencias.get(termo, 0))
    )
    resultado["Número de músicas"] = resultado["Termo_normalizado"].map(
        lambda termo: len(musicas_por_termo.get(termo, set()))
    )

    resultado = resultado.drop(columns=["Termo_normalizado"])
    resultado = resultado.sort_values(
        by=["Frequência", "Número de músicas", "Termo"],
        ascending=[False, False, True],
        kind="stable",
    ).reset_index(drop=True)

    return resultado


def resumir_grupo(
    nome_grupo: str,
    subconjunto: pd.DataFrame,
    frequencias: Counter[str],
    musicas_por_termo: dict[str, set[int]],
) -> dict[str, object]:
    termos = set(subconjunto["Termo_normalizado"])

    frequencia_total = sum(
        frequencias.get(termo, 0)
        for termo in termos
    )

    musicas_do_grupo: set[int] = set()
    for termo in termos:
        musicas_do_grupo.update(
            musicas_por_termo.get(termo, set())
        )

    termos_encontrados = sum(
        1
        for termo in termos
        if frequencias.get(termo, 0) > 0
    )

    return {
        "Grupo": nome_grupo,
        "Frequência total": int(frequencia_total),
        "Número de músicas": len(musicas_do_grupo),
        "Termos na lista": len(termos),
        "Termos encontrados": termos_encontrados,
    }


def montar_resumo_alvos(
    listas: pd.DataFrame,
    frequencias: Counter[str],
    musicas_por_termo: dict[str, set[int]],
) -> pd.DataFrame:
    alvos = listas[
        listas["Tipo"].str.casefold().eq("alvo")
    ].copy()

    origem = alvos["Origem_normalizada"]
    categoria = alvos["Categoria"].str.casefold()

    grupos = [
        (
            "nomes_próprios_masc",
            alvos[origem.eq("nome_próprio_masc")],
        ),
        (
            "nomes_próprios_fem",
            alvos[origem.eq("nome_próprio_fem")],
        ),
        (
            "alvos masculinos sem nomes próprios",
            alvos[
                categoria.eq("masculino")
                & ~origem.isin(
                    {"nome_próprio_masc", "nome_próprio_fem"}
                )
            ],
        ),
        (
            "alvos femininos sem nomes próprios",
            alvos[
                categoria.eq("feminino")
                & ~origem.isin(
                    {"nome_próprio_masc", "nome_próprio_fem"}
                )
            ],
        ),
    ]

    resumo = pd.DataFrame(
        [
            resumir_grupo(
                nome_grupo,
                subconjunto,
                frequencias,
                musicas_por_termo,
            )
            for nome_grupo, subconjunto in grupos
        ]
    )

    resumo = resumo.sort_values(
        by=["Frequência total", "Número de músicas", "Grupo"],
        ascending=[False, False, True],
        kind="stable",
    ).reset_index(drop=True)

    return resumo


# ============================================================
# FORMATAÇÃO DO XLSX
# ============================================================

def formatar_planilha(caminho_saida: Path) -> None:
    fill_cabecalho = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )
    fonte_cabecalho = Font(
        color="FFFFFF",
        bold=True,
    )

    from openpyxl import load_workbook

    workbook = load_workbook(caminho_saida)

    larguras = {
        "Frequencia_alvos": {
            "A": 25,
            "B": 14,
            "C": 18,
            "D": 24,
            "E": 15,
            "F": 20,
        },
        "Frequencia_atributos": {
            "A": 25,
            "B": 14,
            "C": 18,
            "D": 24,
            "E": 15,
            "F": 20,
        },
        "Resumo_alvos": {
            "A": 43,
            "B": 18,
            "C": 20,
            "D": 17,
            "E": 20,
        },
    }

    for nome_aba in workbook.sheetnames:
        aba = workbook[nome_aba]
        aba.freeze_panes = "A2"
        aba.auto_filter.ref = aba.dimensions

        for celula in aba[1]:
            celula.fill = fill_cabecalho
            celula.font = fonte_cabecalho
            celula.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        for coluna, largura in larguras.get(nome_aba, {}).items():
            aba.column_dimensions[coluna].width = largura

        for linha in aba.iter_rows(min_row=2):
            for celula in linha:
                celula.alignment = Alignment(
                    vertical="top",
                    wrap_text=False,
                )

        for coluna in range(5, aba.max_column + 1):
            for linha in range(2, aba.max_row + 1):
                aba.cell(linha, coluna).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

    workbook.save(caminho_saida)


# ============================================================
# EXECUÇÃO
# ============================================================

def main() -> None:
    if not ARQUIVO_CSV.exists():
        raise FileNotFoundError(
            f"CSV não encontrado:\n{ARQUIVO_CSV}"
        )

    if not ARQUIVO_LISTAS.exists():
        raise FileNotFoundError(
            f"XLSX de listas não encontrado:\n{ARQUIVO_LISTAS}"
        )

    print("Lendo as listas de alvos e atributos...")
    listas = carregar_listas(ARQUIVO_LISTAS)

    print("Detectando o formato do CSV...")
    encoding, separador, colunas_csv = detectar_csv(ARQUIVO_CSV)
    coluna_letra = localizar_coluna_letra_musica(colunas_csv)

    print(f"Encoding detectado: {encoding}")
    print(f"Separador detectado: {repr(separador)}")
    print(f"ÚNICA coluna que será lida: {coluna_letra}")

    dados_letras = pd.read_csv(
        ARQUIVO_CSV,
        sep=separador,
        encoding=encoding,
        usecols=[coluna_letra],
        dtype={coluna_letra: "string"},
        keep_default_na=False,
        low_memory=False,
    )

    letras = (
        dados_letras[coluna_letra]
        .fillna("")
        .map(normalizar_texto_busca)
    )

    print(f"Quantidade de letras/músicas lidas: {len(letras):,}")
    print("Contando termos SOMENTE nas letras das músicas...")

    frequencias, musicas_por_termo = contar_nas_letras(
        letras=letras,
        termos=list(listas["Termo_normalizado"]),
    )

    tabela_completa = montar_frequencia_por_termo(
        listas=listas,
        frequencias=frequencias,
        musicas_por_termo=musicas_por_termo,
    )

    aba_alvos = (
        tabela_completa[
            tabela_completa["Tipo"].str.casefold().eq("alvo")
        ]
        .sort_values(
            by=["Frequência", "Número de músicas", "Termo"],
            ascending=[False, False, True],
            kind="stable",
        )
        .reset_index(drop=True)
    )

    aba_atributos = (
        tabela_completa[
            tabela_completa["Tipo"].str.casefold().eq("atributo")
        ]
        .sort_values(
            by=["Frequência", "Número de músicas", "Termo"],
            ascending=[False, False, True],
            kind="stable",
        )
        .reset_index(drop=True)
    )

    aba_resumo = montar_resumo_alvos(
        listas=listas,
        frequencias=frequencias,
        musicas_por_termo=musicas_por_termo,
    )

    print("Gravando o arquivo Excel...")

    with pd.ExcelWriter(
        ARQUIVO_SAIDA,
        engine="openpyxl",
    ) as writer:
        aba_alvos.to_excel(
            writer,
            sheet_name="Frequencia_alvos",
            index=False,
        )
        aba_atributos.to_excel(
            writer,
            sheet_name="Frequencia_atributos",
            index=False,
        )
        aba_resumo.to_excel(
            writer,
            sheet_name="Resumo_alvos",
            index=False,
        )

    formatar_planilha(ARQUIVO_SAIDA)

    print("\nConcluído.")
    print(f"Arquivo gerado:\n{ARQUIVO_SAIDA}")


if __name__ == "__main__":
    main()
